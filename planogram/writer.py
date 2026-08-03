from __future__ import annotations
from io import BytesIO
from typing import List

import pandas as pd
from openpyxl import load_workbook as openpyxl_load
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from planogram.config import (
    OUTPUT_SHEET_PLANOGRAM,
    OUTPUT_SHEET_VALIDATION,
    PLANOGRAM_COLUMNS,
)

# ---------------------------------------------------------------------------
# Styling constants (Home Depot brand colours)
# ---------------------------------------------------------------------------
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HD_ORANGE    = PatternFill(fill_type="solid", fgColor="F96302")
_ERROR_RED    = PatternFill(fill_type="solid", fgColor="CC0000")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, fill: PatternFill = _HD_ORANGE) -> None:
    """Apply bold white text and a coloured fill to the first row of *ws*."""
    for cell in ws[1]:
        cell.font      = _HEADER_FONT
        cell.fill      = fill
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 22


def _auto_width(ws, max_width: int = 60) -> None:
    """Set each column's width to fit its longest value (capped at *max_width*)."""
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(length + 3, max_width)


def _str(val) -> str:
    """Convert a value to string; NaN/None → empty string."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def write_output_bytes(
    input_bytes: bytes,
    planogram_df: pd.DataFrame,
    validation_df: pd.DataFrame,
) -> bytes:
    """Generate the output workbook in memory and return it as bytes.

    The original input workbook is copied first so all source sheets are
    preserved in the output.  The two generated sheets are then appended.

    Parameters
    ----------
    input_bytes:
        Raw bytes of the original uploaded .xlsx file.
    planogram_df:
        Generated planogram DataFrame (from generate_planogram).
    validation_df:
        Validation issues DataFrame (from generate_planogram).

    Returns
    -------
    bytes
        Complete .xlsx file ready for Streamlit's st.download_button.
    """
    # Load the original workbook from bytes (preserves all source sheets)
    wb = openpyxl_load(BytesIO(input_bytes))

    # Remove pre-existing generated sheets so the function is idempotent
    for name in [OUTPUT_SHEET_PLANOGRAM, OUTPUT_SHEET_VALIDATION]:
        if name in wb.sheetnames:
            del wb[name]

    # ── Generated Planogram sheet ───────────────────────────────────────────
    ws_pog = wb.create_sheet(title=OUTPUT_SHEET_PLANOGRAM)
    ws_pog.append(PLANOGRAM_COLUMNS)  # header row

    for _, row in planogram_df.iterrows():
        ws_pog.append([row[col] for col in PLANOGRAM_COLUMNS])

    _style_header(ws_pog, fill=_HD_ORANGE)
    _auto_width(ws_pog)
    ws_pog.freeze_panes = "A2"  # keep header visible while scrolling

    # ── Validation sheet ───────────────────────────────────────────────────
    ws_val = wb.create_sheet(title=OUTPUT_SHEET_VALIDATION)
    val_cols: List[str] = ["Level", "Store", "Bay#", "Message"]
    ws_val.append(val_cols)

    for _, row in validation_df.iterrows():
        ws_val.append([_str(row.get(c, "")) for c in val_cols])

    _style_header(ws_val, fill=_ERROR_RED)
    _auto_width(ws_val)
    ws_val.freeze_panes = "A2"

    # ── Save to BytesIO and return ─────────────────────────────────────────
    output_bio = BytesIO()
    wb.save(output_bio)
    output_bio.seek(0)
    return output_bio.read()
