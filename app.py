"""
Home Depot Planogram Generator — Single-File Streamlit App
===========================================================
All modules are bundled inline so this file can be deployed to
Streamlit Cloud by uploading just this file + requirements.txt.

requirements.txt contents:
    streamlit>=1.28.0
    pandas>=2.0.0
    openpyxl>=3.1.0
"""
from __future__ import annotations

import math
import re
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook as _openpyxl_load
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — PAGE CONFIG  (must be first Streamlit call)
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="HD Planogram Generator",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — MODELS
# ══════════════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class BayRule:
    """Immutable capacity specification for one bay type."""
    display: int
    so:      int
    stock:   int

    @property
    def total(self) -> int:
        return self.display + self.so + self.stock


@dataclass
class SKURecord:
    """One planogram position (a single facing of one physical SKU)."""
    sku:         str
    description: str
    facing:      int
    sku_type:    str   # "Display" | "SO" | "Stock"
    match_name:  str = ""  # clean product Name, used only for exact conflict matching


@dataclass
class SORecord:
    """One entry in the global Special Order ranked list."""
    rank:        int
    category:    str   # "Stone" | "Wood" (or any string from the sheet)
    sku:         str
    description: str
    omsid:       str
    cf:          float = 0.0  # Color Flow rank — used to order the final selection
    name:        str = ""     # clean product Name, used only for exact conflict matching


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

BAY_RULES: Dict[str, BayRule] = {
    "99":  BayRule(display=6, so=2, stock=6),
    "99C": BayRule(display=4, so=2, stock=4),
    "87":  BayRule(display=4, so=1, stock=4),
    "75":  BayRule(display=4, so=1, stock=4),
    "51":  BayRule(display=2, so=1, stock=2),
}

# Notes rules:
#   trigger_keywords — if ANY of these appear in the Notes cell, the rule fires.
#   sku_keywords     — SKUs whose description contains ANY of these are moved to tail.
#   "baja" matches: BAJA, BAJA BEIGE, BAJA WHITE, BAJABEIGE, BAJA DARK GREY ...
#   "vigo" matches: VIGO, VIGO GREY, VIGO BEIGE, VIGOBEIGE ...
NOTES_RULES: List[Dict[str, Any]] = [
    {
        "trigger_keywords": ["baja", "vigo"],
        "sku_keywords":     ["baja", "vigo"],
    },
]

# Sheet names (matched case-insensitively at load time)
SHEET_STORE_LIST    = "Store List"
SHEET_STOCK_DISPLAY = "Stock SKUs and Displays"
SHEET_SO            = "Special Order Boards"

# Output sheet names
OUTPUT_SHEET_PLANOGRAM  = "Generated Planogram"
OUTPUT_SHEET_VALIDATION = "Validation"
OUTPUT_SHEET_STORE_POG  = "Store List (POG)"
OUTPUT_SHEET_AMT        = "AMT"

# Output column order
PLANOGRAM_COLUMNS: List[str] = [
    "Store", "Bay#", "Bay Size", "Shelf", "Position",
    "SKU", "SKU Type", "SKU Description", "Facing",
]

# Shelf numbering
SHELF_DISPLAY      = 1
SHELF_SO           = 2
SHELF_STOCK_FIRST  = 3
SHELF_STOCK_SECOND = 4

# LFT sentinel values — these mean "no LFT for this store"
LFT_IGNORE_VALUES: frozenset = frozenset(["-", "", "none", "null", "n/a", "na"])

# Valid Facing values — anything outside this set is clamped to 1
VALID_FACINGS: frozenset = frozenset([1, 2])


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — LOGGER
# ══════════════════════════════════════════════════════════════════════════════

class PlanogramLogger:
    """Collects validation issues. Create a fresh instance per generation run."""

    def __init__(self) -> None:
        self.issues: List[Dict[str, Any]] = []
        self._seen: set = set()

    def _append(self, level: str, msg: str, store: Any = "", bay_num: Any = "") -> None:
        key = (level, str(store), str(bay_num), msg)
        if key in self._seen:
            return
        self._seen.add(key)
        self.issues.append({"Level": level, "Store": str(store), "Bay#": str(bay_num), "Message": msg})

    def warning(self, msg: str, store: Any = "", bay_num: Any = "") -> None:
        self._append("WARNING", msg, store, bay_num)

    def error(self, msg: str, store: Any = "", bay_num: Any = "") -> None:
        self._append("ERROR", msg, store, bay_num)

    def info(self, msg: str) -> None:
        pass  # INFO messages are not surfaced in the Validation sheet


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — POG / LFT STRING PARSER
# ══════════════════════════════════════════════════════════════════════════════

# Dash after "Bay" is OPTIONAL to handle:
#   "8 Bay 99,99,..."          (no dash)
#   "1 Bay 99C - LFT - Full"   (no dash, single token)
#   "6 Bay-99,99,..."          (dash attached)
#   "8 Bay - 99 - description" (standard)
_RE_LEADING_COUNT = re.compile(r"^\s*(\d+)\s+bay\s*[-\u2013]?\s*", re.IGNORECASE)

# Matches a complete clean token (no trailing description)
_RE_VALID_TOKEN = re.compile(r"^[A-Za-z0-9]+(?:-[A-Za-z0-9]+)?$")

# Matches a multiplier pair: "<count>-<size>"  e.g. "8-99", "2-51"
_RE_MULTIPLIER_TOKEN = re.compile(r"^(\d+)-([A-Za-z0-9]+)$")

# Extracts the leading bay-size token from a string with trailing description.
# Handles plain sizes ("99", "99C") AND multiplier pairs ("1-87", "3-99").
# Safety: requires DIGITS after the dash, so "99-Stone" → "99" (not "99-Stone").
_RE_BAYSIZE_PREFIX = re.compile(r"^(\d+[A-Za-z]{0,2}(?:-\d+[A-Za-z]{0,2})?)")


def parse_pog_string(
    pog: str,
    store: Any = "",
    logger: PlanogramLogger | None = None,
) -> List[str]:
    """Parse a POG/LFT string into an ordered list of bay-size tokens.

    FORMAT A: "10 Bay - 8-99,2-51"         → ["99"]*8 + ["51"]*2
    FORMAT B: "7 BAY - 99,99,99,99,99,87"  → explicit list
    FORMAT C: "8 Bay - 99"                  → ["99"]*8
    """
    if not isinstance(pog, str) or not pog.strip():
        return []

    pog = pog.strip()
    m_count = _RE_LEADING_COUNT.match(pog)
    if not m_count:
        if logger:
            logger.warning(f"Cannot parse bay count from POG: '{pog}'", store=store)
        return []

    declared_count: int = int(m_count.group(1))
    remainder: str = pog[m_count.end():].strip()

    raw_tokens = [t.strip() for t in remainder.split(",")]
    bay_tokens_raw: List[str] = []

    for tok in raw_tokens:
        if _RE_VALID_TOKEN.match(tok):
            bay_tokens_raw.append(tok)
        else:
            m_prefix = _RE_BAYSIZE_PREFIX.match(tok)
            if m_prefix:
                bay_tokens_raw.append(m_prefix.group(1))
            break  # description text signals end of bay tokens

    if not bay_tokens_raw:
        if logger:
            logger.warning(f"No bay tokens found in POG: '{pog}'", store=store)
        return []

    all_multiplier = all(_RE_MULTIPLIER_TOKEN.match(tok) for tok in bay_tokens_raw)

    if all_multiplier:
        expanded: List[str] = []
        for tok in bay_tokens_raw:
            m = _RE_MULTIPLIER_TOKEN.match(tok)
            expanded.extend([m.group(2).upper()] * int(m.group(1)))
    elif len(bay_tokens_raw) == 1:
        expanded = [bay_tokens_raw[0].upper()] * declared_count
    else:
        expanded = [tok.upper() for tok in bay_tokens_raw]

    if len(expanded) != declared_count and logger:
        logger.warning(
            f"Parsed {len(expanded)} bay(s) but header declares {declared_count}: '{pog}'",
            store=store,
        )

    return expanded


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — WORKBOOK LOADER
# ══════════════════════════════════════════════════════════════════════════════

def _str(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def _clean_store_id(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    try:
        f_val = float(s)
        if f_val.is_integer():
            return str(int(f_val))
    except (ValueError, AttributeError):
        pass
    return s


def _norm(name: str) -> str:
    return str(name).strip().lower()


def _find_col(
    df: pd.DataFrame,
    candidates: List[str],
    fallback_pos: int,
    sheet_name: str,
    logger: PlanogramLogger | None = None,
) -> str:
    cols = list(df.columns)
    norm_map = {_norm(c): c for c in cols}

    for cand in candidates:
        if _norm(cand) in norm_map:
            return norm_map[_norm(cand)]

    if fallback_pos < len(cols):
        actual = cols[fallback_pos]
        if logger:
            logger.info(f"[{sheet_name}] Header '{candidates[0]}' not found; using position {fallback_pos} ('{actual}').")
        return actual

    raise ValueError(
        f"[{sheet_name}] Cannot find column '{candidates[0]}' by name or "
        f"position {fallback_pos}. Available columns: {cols}"
    )


@dataclass
class WorkbookData:
    store_list:     pd.DataFrame
    stock_display:  pd.DataFrame
    special_orders: pd.DataFrame
    cols_sl: Dict[str, str] = field(default_factory=dict)
    cols_sd: Dict[str, str] = field(default_factory=dict)
    cols_so: Dict[str, str] = field(default_factory=dict)


def load_workbook_from_bytes(
    file_bytes: bytes,
    logger: PlanogramLogger | None = None,
) -> WorkbookData:
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes), engine="openpyxl")
    except Exception as exc:
        raise RuntimeError(f"Cannot open workbook: {exc}") from exc

    def _get_sheet(desired: str) -> str:
        for name in xl.sheet_names:
            if name.strip().lower() == desired.strip().lower():
                return name
        raise ValueError(f"Sheet '{desired}' not found. Available: {xl.sheet_names}")

    s_sl = _get_sheet(SHEET_STORE_LIST)
    s_sd = _get_sheet(SHEET_STOCK_DISPLAY)
    s_so = _get_sheet(SHEET_SO)

    df_sl = xl.parse(s_sl, dtype=str)
    df_sd = xl.parse(s_sd, dtype=str)
    df_so = xl.parse(s_so, dtype=str)

    for df in (df_sl, df_sd, df_so):
        df.columns = [str(c).strip() for c in df.columns]
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)

    cols_sl = {
        "store": _find_col(df_sl, ["Store"], 0, s_sl, logger),
        "pog":   _find_col(df_sl, ["Current Store POG", "Store POG", "Current Pog", "POG"], 1, s_sl, logger),
        "lft":   _find_col(df_sl, ["Current LFT", "LFT", "Current Lft"], 2, s_sl, logger),
        "notes": _find_col(df_sl, ["Notes", "Note"], 3, s_sl, logger),
    }

    sd_cols = list(df_sd.columns)
    cols_sd = {
        "store":      _find_col(df_sd, ["Store"], 0, s_sd, logger),
        "stock_sku":  _find_col(df_sd, ["Stock SKU", "Stock Sku", "SKU"], 1, s_sd, logger),
        "stock_name": _find_col(df_sd, ["Name", "Stock Name", "Product Name"], 2, s_sd, logger),
        "stock_desc": _find_col(df_sd, ["Stock Description", "Stock Desc", "Description"], 3, s_sd, logger),
        "stock_face": _find_col(df_sd, ["Facings", "Facing", "Stock Facings", "Stock Facing"], 5, s_sd, logger),
        "disp_sku":   _find_col(df_sd, ["Display SKU", "Display Sku"], 6, s_sd, logger),
        "disp_desc":  _find_col(df_sd, ["Display Description", "Display Desc"], 7, s_sd, logger),
        "disp_face":  _find_col(df_sd, ["Display Facing", "Display Facings", "Display", "Facings.1", "Facing.1"], 9 if len(sd_cols) <= 11 else 10, s_sd, logger),
        "cf":         _find_col(df_sd, ["Final CF", "CF"], len(sd_cols) - 1, s_sd, logger),
    }

    # Sort Stock & Display sheet by CF (ascending) so allocation is always in
    # the correct order even if the sheet was saved unsorted.
    cf_col = cols_sd["cf"]
    df_sd[cf_col] = pd.to_numeric(df_sd[cf_col], errors="coerce")
    df_sd.sort_values(by=cf_col, ascending=True, na_position="last", inplace=True)
    df_sd.reset_index(drop=True, inplace=True)

    cols_so = {
        "combined_rank": _find_col(df_so, ["Combined Rank", "Rank", "Combined rank"], 0, s_so, logger),
        "category":      _find_col(df_so, ["Category", "category"], 1, s_so, logger),
        "sku":           _find_col(df_so, ["SKU", "Sku"], 2, s_so, logger),
        "name":          _find_col(df_so, ["Name", "Product Name"], 3, s_so, logger),
        "sku_desc":      _find_col(df_so, ["SKU Description", "Description", "Sku Description"], 4, s_so, logger),
        "omsid":         _find_col(df_so, ["OMSID", "OMS ID", "Omsid"], 5, s_so, logger),
        "cf":            _find_col(df_so, ["CF", "Color Flow", "CF Rank"], 6, s_so, logger),
    }

    return WorkbookData(
        store_list=df_sl,
        stock_display=df_sd,
        special_orders=df_so,
        cols_sl=cols_sl,
        cols_sd=cols_sd,
        cols_so=cols_so,
    )


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — BAY ALLOCATOR
# ══════════════════════════════════════════════════════════════════════════════

def _parse_facing(raw: Any, store: Any, sku: Any, logger: PlanogramLogger) -> int:
    """Always returns 1 or 2. Values outside {1,2} are clamped to 1."""
    try:
        val = int(float(str(raw).strip()))
    except (ValueError, TypeError):
        logger.warning(f"Invalid Facing '{raw}' for SKU {sku} — defaulting to 1.", store=store)
        return 1
    if val not in VALID_FACINGS:
        logger.warning(f"Facing '{raw}' for SKU {sku} not in {{1,2}} — defaulting to 1.", store=store)
        return 1
    return val


def build_display_index(df: pd.DataFrame, cols: Dict[str, str], logger: PlanogramLogger) -> Dict[str, List[SKURecord]]:
    index: Dict[str, List[SKURecord]] = defaultdict(list)
    for _, row in df.iterrows():
        store  = _clean_store_id(row[cols["store"]])
        sku    = _str(row[cols["disp_sku"]])
        desc   = _str(row[cols["disp_desc"]])
        disp_f  = _parse_facing(row[cols["disp_face"]], store, sku, logger)
        stock_f = _parse_facing(row[cols["stock_face"]], store, sku, logger)
        # Display Board facing uses max(Display Facing, Stock Facing)
        facing  = max(disp_f, stock_f)
        if not store or not sku:
            continue
        index[store].append(SKURecord(sku=sku, description=desc, facing=facing, sku_type="Display"))
    return dict(index)


def build_stock_index(df: pd.DataFrame, cols: Dict[str, str], logger: PlanogramLogger) -> Dict[str, List[SKURecord]]:
    index: Dict[str, List[SKURecord]] = defaultdict(list)
    for _, row in df.iterrows():
        store  = _clean_store_id(row[cols["store"]])
        sku    = _str(row[cols["stock_sku"]])
        desc   = _str(row[cols["stock_desc"]])
        name   = _str(row[cols["stock_name"]]) if "stock_name" in cols else ""
        facing = _parse_facing(row[cols["stock_face"]], store, sku, logger)
        if not store or not sku:
            continue
        index[store].append(SKURecord(sku=sku, description=desc, facing=facing, sku_type="Stock", match_name=name))
    return dict(index)


def build_so_global_list(df: pd.DataFrame, cols: Dict[str, str]) -> List[SORecord]:
    """Load the global SO ranked list (sorted by Combined Rank)."""
    records: List[SORecord] = []
    for _, row in df.iterrows():
        try:
            rank = int(float(str(row[cols["combined_rank"]]).strip()))
        except (ValueError, TypeError):
            continue
        category = _str(row[cols["category"]])
        sku      = _str(row[cols["sku"]])
        desc     = _str(row[cols["sku_desc"]])
        name     = _str(row[cols["name"]]) if "name" in cols else ""
        omsid    = _str(row[cols["omsid"]])
        # CF: parse to float; #N/A / blank / non-numeric → sort to end
        try:
            cf = float(str(row[cols["cf"]]).strip())
        except (ValueError, TypeError, KeyError):
            cf = float("inf")
        if not sku:
            continue
        records.append(SORecord(rank=rank, category=category, sku=sku, description=desc, omsid=omsid, cf=cf, name=name))
    records.sort(key=lambda r: r.rank)
    return records



# ---------------------------------------------------------------------------
# SO conflict-avoidance helpers
# ---------------------------------------------------------------------------

def _norm_product_name(name: str) -> str:
    """Normalize a product Name for exact, case-insensitive comparison.

    Collapses repeated whitespace and upper-cases. No fuzzy matching, no
    keyword extraction — this is a direct string comparison on the clean
    product Name field (e.g. 'BIANCO DOLOMI'), not the full SKU description.
    """
    if not isinstance(name, str) or not name.strip():
        return ""
    return re.sub(r"\s+", " ", name.strip()).upper()


def _build_conflict_set(stock_names: List[str]) -> set:
    """Build the set of normalized stock product Names already in a store.

    An SO item conflicts only if its product Name is an exact (case-
    insensitive) match to one of these — e.g. stock 'CRYSTAL BIANCO' and SO
    'BIANCO DOLOMI' are distinct products and do NOT conflict, but stock
    'BIANCO DOLOMI' and SO 'BIANCO DOLOMI' do.
    """
    return {_norm_product_name(n) for n in stock_names if _norm_product_name(n)}


def _is_so_conflict(so_name: str, conflict_entries: set) -> bool:
    """Return True if so_name exactly matches (case-insensitive) a stock product Name."""
    norm = _norm_product_name(so_name)
    if not norm:
        return False
    return norm in conflict_entries


def select_so_for_store(
    so_global_list: List[SORecord],
    conflict_entries: set,
    capacity: int,
    store: Any,
    logger: PlanogramLogger,
) -> List[SKURecord]:
    """Pick SO SKUs for one store from the global ranked list.

    Walk in Combined Rank order. For every position:
    - Available & no conflict  → take it directly.
    - Already used as a prior replacement  → this slot's category must still be
      filled; find the next non-used, non-conflicting item of the SAME category.
    - Conflict with stock      → same as above; find next of the same category.

    This guarantees the category pattern from the global list (e.g. every 5th
    position is Wood) is preserved even when replacements cascade forward.
    """
    # Pre-index items by category for same-category forward search
    cat_lists: Dict[str, List[SORecord]] = defaultdict(list)
    for rec in so_global_list:          # already sorted by rank
        cat_lists[rec.category.lower()].append(rec)

    def _next_in_category(cat: str, after_rank: int) -> SORecord | None:
        """Return the first non-used, non-conflicting SO of `cat` with rank > after_rank."""
        for candidate in cat_lists[cat]:
            if candidate.rank <= after_rank:
                continue
            if candidate.rank in used_ranks:
                continue
            if not _is_so_conflict(candidate.name, conflict_entries):
                return candidate
        return None

    used_ranks: set = set()
    # Track (cf_value, category, SKURecord) so we can sort by CF per category
    selected_triples: List[Tuple[float, str, SKURecord]] = []

    def _sel_count() -> int:
        return len(selected_triples)

    for rec in so_global_list:
        if _sel_count() >= capacity:
            break

        already_used = rec.rank in used_ranks
        has_conflict = _is_so_conflict(rec.name, conflict_entries)

        if not already_used and not has_conflict:
            # Happy path — take this item directly.
            used_ranks.add(rec.rank)
            selected_triples.append((rec.cf, rec.category.lower(), SKURecord(
                sku=rec.sku, description=rec.description,
                facing=1, sku_type="SO",
            )))
        else:
            cat = rec.category.lower()
            replacement = _next_in_category(cat, rec.rank)

            if replacement:
                used_ranks.add(replacement.rank)
                selected_triples.append((replacement.cf, replacement.category.lower(), SKURecord(
                    sku=replacement.sku, description=replacement.description,
                    facing=1, sku_type="SO",
                )))
                if has_conflict and not already_used:
                    logger.info(
                        f"Store {store}: replaced conflicting SO rank {rec.rank} "
                        f"({rec.category} '{rec.sku}') with rank {replacement.rank} "
                        f"('{replacement.sku}')."
                    )
            else:
                # No further item of this category is available.
                if not already_used:
                    logger.warning(
                        f"No non-conflicting {rec.category} SO available to replace "
                        f"rank {rec.rank} ('{rec.sku}') for store {store}.",
                        store=store,
                    )

    # ── Final step: sort by CF within each category, then interleave ──────────
    # Stone and Wood are sorted by CF independently, then assembled into the
    # 4-Stone / 1-Wood shelf pattern:
    #   positions 1-4 = Stone (lowest CF first)
    #   position  5   = Wood  (lowest CF first)
    #   positions 6-9 = Stone, position 10 = Wood, …
    # If one category runs out, the other fills in as fallback.
    stone_items = [rec for _, cf, rec in sorted(
        [(0, cf, rec) for cf, cat, rec in selected_triples if cat != "wood"],
        key=lambda t: t[1],
    )]
    wood_items = [rec for _, cf, rec in sorted(
        [(0, cf, rec) for cf, cat, rec in selected_triples if cat == "wood"],
        key=lambda t: t[1],
    )]

    result:     List[SKURecord] = []
    stone_idx = 0
    wood_idx  = 0
    for pos in range(1, len(selected_triples) + 1):
        if pos % 5 == 0:                        # Wood slot
            if wood_idx < len(wood_items):
                result.append(wood_items[wood_idx]); wood_idx += 1
            elif stone_idx < len(stone_items):  # fallback: no more wood
                result.append(stone_items[stone_idx]); stone_idx += 1
        else:                                   # Stone slot (positions 1-4, 6-9, …)
            if stone_idx < len(stone_items):
                result.append(stone_items[stone_idx]); stone_idx += 1
            elif wood_idx < len(wood_items):    # fallback: no more stone
                result.append(wood_items[wood_idx]); wood_idx += 1
    return result



def expand_facing(records: List[SKURecord]) -> List[SKURecord]:
    expanded: List[SKURecord] = []
    for rec in records:
        for _ in range(max(1, rec.facing)):
            expanded.append(rec)
    return expanded


def apply_notes_rules(records: List[SKURecord], notes: str) -> List[SKURecord]:
    """Move matching SKUs to tail via stable partition (no re-sort)."""
    if not notes or not isinstance(notes, str):
        return records

    notes_lower = notes.strip().lower()
    result = list(records)

    for rule in NOTES_RULES:
        triggers = list(rule.get("trigger_keywords", []))
        if "trigger" in rule:
            triggers.append(rule["trigger"])

        if not any(tr.lower() in notes_lower for tr in triggers if tr):
            continue

        sku_kws = [kw.lower() for kw in rule.get("sku_keywords", rule.get("keywords", [])) if kw]
        if not sku_kws:
            continue

        head: List[SKURecord] = []
        tail: List[SKURecord] = []
        for rec in result:
            if any(kw in rec.description.lower() for kw in sku_kws):
                tail.append(rec)
            else:
                head.append(rec)
        result = head + tail

    return result


def _consume(
    pool: List[SKURecord], pointer: List[int], count: int,
    store: Any, bay_num: int, sku_type: str, logger: PlanogramLogger,
) -> List[SKURecord]:
    start = pointer[0]
    end   = start + count
    taken = pool[start:end]
    if len(taken) < count:
        logger.warning(
            f"Not enough {sku_type} SKUs for Bay {bay_num}: need {count}, have {len(taken)} remaining.",
            store=store, bay_num=bay_num,
        )
    pointer[0] = min(end, len(pool))
    return taken


def allocate_bay(
    store: Any, bay_num: int, bay_size: str, rule: BayRule,
    disp_pool: List[SKURecord], so_pool: List[SKURecord], stock_pool: List[SKURecord],
    disp_ptr: List[int], so_ptr: List[int], stock_ptr: List[int],
    logger: PlanogramLogger,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []

    def _row(shelf: int, pos: int, rec: SKURecord) -> Dict[str, Any]:
        return {
            "Store": store, "Bay#": bay_num, "Bay Size": bay_size,
            "Shelf": shelf, "Position": pos,
            "SKU": rec.sku, "SKU Type": rec.sku_type,
            "SKU Description": rec.description, "Facing": rec.facing,
        }

    for pos, rec in enumerate(_consume(disp_pool, disp_ptr, rule.display, store, bay_num, "Display", logger), start=1):
        rows.append(_row(SHELF_DISPLAY, pos, rec))

    for pos, rec in enumerate(_consume(so_pool, so_ptr, rule.so, store, bay_num, "SO", logger), start=1):
        rows.append(_row(SHELF_SO, pos, rec))

    shelf3_count  = math.ceil(rule.stock / 2)
    stock_records = _consume(stock_pool, stock_ptr, rule.stock, store, bay_num, "Stock", logger)
    for pos, rec in enumerate(stock_records[:shelf3_count], start=1):
        rows.append(_row(SHELF_STOCK_FIRST, pos, rec))
    for pos, rec in enumerate(stock_records[shelf3_count:], start=1):
        rows.append(_row(SHELF_STOCK_SECOND, pos, rec))

    return rows


def allocate_store(
    store: str, notes: str, bay_list: List[str],
    raw_display: List[SKURecord],
    raw_so: List[SKURecord],      # pre-selected & conflict-free for this store
    raw_stock: List[SKURecord],
    logger: PlanogramLogger,
) -> List[Dict[str, Any]]:
    disp_pool  = expand_facing(apply_notes_rules(raw_display, notes))
    so_pool    = list(raw_so)     # SO order is fixed by global rank; no notes rules
    stock_pool = expand_facing(apply_notes_rules(raw_stock,   notes))

    disp_ptr  = [0]
    so_ptr    = [0]
    stock_ptr = [0]
    all_rows: List[Dict[str, Any]] = []

    for bay_num, bay_size in enumerate(bay_list, start=1):
        rule = BAY_RULES.get(bay_size)
        if rule is None:
            logger.warning(f"Unsupported Bay Size '{bay_size}' in Bay {bay_num} — skipping.", store=store, bay_num=bay_num)
            continue
        all_rows.extend(
            allocate_bay(
                store=store, bay_num=bay_num, bay_size=bay_size, rule=rule,
                disp_pool=disp_pool, so_pool=so_pool, stock_pool=stock_pool,
                disp_ptr=disp_ptr, so_ptr=so_ptr, stock_ptr=stock_ptr,
                logger=logger,
            )
        )

    for label, pool, ptr in [("Display", disp_pool, disp_ptr), ("SO", so_pool, so_ptr), ("Stock", stock_pool, stock_ptr)]:
        remaining = len(pool) - ptr[0]
        if remaining > 0:
            logger.warning(f"{remaining} {label} position(s) unused after all bays allocated.", store=store)

    return all_rows


def generate_planogram(
    wb_data: WorkbookData,
    logger: PlanogramLogger,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    disp_index     = build_display_index(wb_data.stock_display, wb_data.cols_sd, logger)
    stock_index    = build_stock_index(wb_data.stock_display,   wb_data.cols_sd, logger)
    so_global_list = build_so_global_list(wb_data.special_orders, wb_data.cols_so)

    store_col    = wb_data.cols_sl["store"]
    store_series = wb_data.store_list[store_col].apply(_clean_store_id)
    for dup in store_series[store_series.duplicated(keep=False)].unique():
        if dup:
            logger.warning(f"Duplicate store '{dup}' in Store List — all rows processed.")

    all_output_rows: List[Dict[str, Any]] = []
    seen_stores: set = set()

    for _, row in wb_data.store_list.iterrows():
        store = _clean_store_id(row[wb_data.cols_sl["store"]])
        if not store:
            logger.warning("Store column is empty for a row — skipping.")
            continue

        pog_raw = _str(row[wb_data.cols_sl["pog"]])
        lft_raw = _str(row[wb_data.cols_sl["lft"]])
        notes   = _str(row[wb_data.cols_sl["notes"]])

        if store in seen_stores:
            logger.warning(f"Store {store} appears more than once — processing duplicate.", store=store)
        seen_stores.add(store)

        pog_bays = parse_pog_string(pog_raw, store=store, logger=logger)
        if not pog_bays:
            logger.warning(f"Malformed/empty POG '{pog_raw}' — skipping store.", store=store)
            continue

        lft_bays: List[str] = []
        if lft_raw.lower() not in LFT_IGNORE_VALUES:
            lft_bays = parse_pog_string(lft_raw, store=store, logger=logger)
            if not lft_bays:
                logger.warning(f"Malformed LFT '{lft_raw}' — ignoring LFT.", store=store)

        bay_list: List[str] = lft_bays + pog_bays  # LFT FIRST, then POG

        raw_display = disp_index.get(store, [])
        raw_stock   = stock_index.get(store, [])

        if not raw_display:
            logger.warning(f"No Display SKUs found for store {store}.", store=store)
        if not raw_stock:
            logger.warning(f"No Stock SKUs found for store {store}.", store=store)

        # ── SO: select from global ranked list with conflict avoidance ──────
        # Capacity = sum of SO slots across all valid bays
        so_capacity = sum(
            BAY_RULES[b].so for b in bay_list if b in BAY_RULES
        )
        # Conflict set = exact product Names already present in this store's stock
        stock_names       = [rec.match_name for rec in raw_stock]
        conflict_entries = _build_conflict_set(stock_names)
        raw_so = select_so_for_store(
            so_global_list, conflict_entries, so_capacity, store, logger
        )
        if not raw_so:
            logger.warning(f"No SO SKUs could be selected for store {store}.", store=store)
        # ────────────────────────────────────────────────────────────────────

        try:
            store_rows = allocate_store(
                store=store, notes=notes, bay_list=bay_list,
                raw_display=raw_display, raw_so=raw_so, raw_stock=raw_stock,
                logger=logger,
            )
        except Exception as exc:
            logger.error(f"Unexpected error allocating store {store}: {exc}", store=store)
            continue

        all_output_rows.extend(store_rows)

    planogram_df = pd.DataFrame(all_output_rows, columns=PLANOGRAM_COLUMNS)
    validation_df = (
        pd.DataFrame(logger.issues)
        if logger.issues
        else pd.DataFrame(columns=["Level", "Store", "Bay#", "Message"])
    )
    return planogram_df, validation_df


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — STORE LIST WITH POG NAME & AMT REPORT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — STORE LIST & AMT EXECUTIVE DASHBOARD BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def _make_pog_name(pog_raw: str, lft_raw: str) -> str:
    """Compute the POG Name label for one store.

    Combines LFT bays (if any) + POG bays into one ordered list, then
    formats as:  '{n} Bay - Reflow Stores ({bay1},{bay2},...)'.
    """
    lft_bays: List[str] = []
    if _str(lft_raw).lower() not in LFT_IGNORE_VALUES:
        lft_bays = parse_pog_string(_str(lft_raw), logger=None) or []
    pog_bays = parse_pog_string(_str(pog_raw), logger=None) or []
    all_bays = lft_bays + pog_bays
    if not all_bays:
        return ""
    return f"{len(all_bays)} Bay - Reflow Stores ({','.join(all_bays)})"


# Canonical display order for bay sizes inside Final Set Name.
# Lower number = shown first in the compact label.
_FINAL_SIZE_ORDER: Dict[str, int] = {"99": 0, "99C": 1, "87/75": 2, "51": 3}


def _normalize_bay_size(raw: str) -> str:
    """Normalise a raw bay-size token: strip whitespace, upper-case, merge 87 and 75."""
    s = raw.strip().upper()
    if s in ("87", "75"):
        return "87/75"
    return s


def _make_final_set_name(pog_name: str) -> str:
    """Compact a POG Name into its Final Set Name.

    Examples
    --------
    "10 Bay - Reflow Stores (99,99,99,99,99,99,99,99,99,99)"
        → "10 Bay - Reflow Stores(99 - 10)"

    "10 Bay - Reflow Stores (99,99,99,99,99,99,99,99,51,51)"
        → "10 Bay - Reflow Stores(99 - 8, 51 - 2)"

    "10 Bay - Reflow Stores (99,99,99,99,75,99,99,99,99,99)"
        → "10 Bay - Reflow Stores(99 - 9, 87/75 - 1)"

    Rules
    -----
    • 87 and 75 are merged into the single label "87/75".
    • Sizes are COUNTED across the whole bay list (position-independent).
    • Output order: 99 → 99C → 87/75 → 51 → anything else (alphabetical).
    • If the pog_name cannot be parsed, it is returned unchanged.
    """
    if not pog_name:
        return pog_name

    # ── Extract bay count from the prefix ─────────────────────────────
    bay_count_m = re.match(r"^\s*(\d+)\s+Bay", pog_name, re.IGNORECASE)
    # ── Extract the comma-separated list inside the last parentheses ───
    paren_m = re.search(r"\(([^)]+)\)\s*$", pog_name)
    if not bay_count_m or not paren_m:
        return pog_name  # can't parse → leave as-is

    bay_count = int(bay_count_m.group(1))
    raw_sizes  = [s.strip() for s in paren_m.group(1).split(",") if s.strip()]
    if not raw_sizes:
        return pog_name

    # Normalise and count
    counts: Dict[str, int] = {}
    for rs in raw_sizes:
        label = _normalize_bay_size(rs)
        counts[label] = counts.get(label, 0) + 1

    # Sort by canonical order, then alphabetically for unknowns
    sorted_labels = sorted(
        counts.keys(),
        key=lambda s: (_FINAL_SIZE_ORDER.get(s, 99), s),
    )

    parts = ", ".join(f"{label} - {counts[label]}" for label in sorted_labels)
    return f"{bay_count} Bay - Reflow Stores({parts})"


def build_store_list_with_pog_name(wb_data: "WorkbookData") -> pd.DataFrame:
    """Return a copy of the Store List sheet with 'POG Name' inserted as Column B."""
    df = wb_data.store_list.copy()
    cols = wb_data.cols_sl
    pog_col = cols.get("pog", "")
    lft_col = cols.get("lft", "")

    pog_names: List[str] = []
    for _, row in df.iterrows():
        pog_raw = _str(row[pog_col]) if pog_col and pog_col in df.columns else ""
        lft_raw = _str(row[lft_col]) if lft_col and lft_col in df.columns else ""
        pog_names.append(_make_pog_name(pog_raw, lft_raw))

    # Compact form: "10 Bay - Reflow Stores(99 - 9, 87/75 - 1)" etc.
    final_set_names: List[str] = [_make_final_set_name(p) for p in pog_names]

    # Insert right after Store (column index 1): POG Name at B, Final Set Name at C
    insert_pos = 1 if len(df.columns) > 1 else len(df.columns)
    df.insert(insert_pos,     "POG Name",       pog_names)
    df.insert(insert_pos + 1, "Final Set Name", final_set_names)
    return df


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 9 — EXCEL WRITER & SHEET STYLES
# ══════════════════════════════════════════════════════════════════════════════

_NAVY_FILL         = PatternFill(fill_type="solid", fgColor="1F4E79")
_LIGHT_GREEN_FILL  = PatternFill(fill_type="solid", fgColor="E2EFDA")
_HD_ORANGE         = PatternFill(fill_type="solid", fgColor="F96302")
_ERROR_RED         = PatternFill(fill_type="solid", fgColor="CC0000")

_TITLE_FONT        = Font(name="Calibri", size=13, bold=True, color="1F4E79")
_SECTION_FONT      = Font(name="Calibri", size=11, bold=True, color="1F4E79")
_WHITE_FONT_BOLD   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_KPI_LABEL_FONT    = Font(name="Calibri", size=10, bold=False, color="000000")
_KPI_VAL_FONT      = Font(name="Calibri", size=10, bold=True, color="000000")
_DATA_FONT         = Font(name="Calibri", size=10, bold=False, color="000000")
_DATA_FONT_BOLD    = Font(name="Calibri", size=10, bold=True, color="000000")

_ALIGN_LEFT        = Alignment(horizontal="left", vertical="center")
_ALIGN_CENTER      = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT       = Alignment(horizontal="right", vertical="center")
_HEADER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)

_THIN_SIDE         = Side(border_style="thin", color="D9D9D9")
_CELL_BORDER       = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def _style_header(ws, fill: PatternFill = _NAVY_FILL) -> None:
    for cell in ws[1]:
        cell.font      = _WHITE_FONT_BOLD
        cell.fill      = fill
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 24


def _auto_width(ws, max_width: int = 70) -> None:
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 10), max_width)


def _write_df_to_sheet(
    wb,
    sheet_title: str,
    df: pd.DataFrame,
    fill: PatternFill = _NAVY_FILL,
) -> None:
    """Create (or replace) a sheet in wb and write df into it with header styling."""
    if sheet_title in wb.sheetnames:
        del wb[sheet_title]
    ws = wb.create_sheet(title=sheet_title)
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([_str(row[c]) if not isinstance(row[c], (int, float)) else row[c] for c in df.columns])
    _style_header(ws, fill=fill)
    _auto_width(ws)
    ws.freeze_panes = "A2"


def _write_amt_sheet(
    wb,
    store_list_df: pd.DataFrame,
    planogram_df: pd.DataFrame,
    wb_data: "WorkbookData",
) -> None:
    """Generate the executive AMT dashboard tab with KPIs, Shared Groups, and Detail."""
    if "AMT" in wb.sheetnames:
        del wb["AMT"]
    ws = wb.create_sheet(title="AMT")

    store_col = wb_data.cols_sl.get("store", "")
    pog_col   = wb_data.cols_sl.get("pog", "")
    lft_col   = wb_data.cols_sl.get("lft", "")

    # 1. Build per-store stock SKU sequences from planogram_df
    stock_sku_seq: Dict[str, List[str]] = {}
    if not planogram_df.empty:
        stock_rows = planogram_df[planogram_df["SKU Type"] == "Stock"].copy()
        for store_id, grp in stock_rows.groupby("Store", sort=False):
            stock_sku_seq[str(store_id)] = grp["SKU"].dropna().astype(str).tolist()

    # 2. Pre-extract store info from store_list_df in original order
    store_info_list: List[Dict[str, Any]] = []
    for _, row in store_list_df.iterrows():
        sid = _clean_store_id(row[store_col]) if store_col and store_col in row else ""
        if not sid:
            continue
        pog_raw = _str(row[pog_col]) if pog_col and pog_col in row else ""
        lft_raw = _str(row[lft_col]) if lft_col and lft_col in row else ""

        # Total bay count
        lft_bays = parse_pog_string(lft_raw, logger=None) if lft_raw.lower() not in LFT_IGNORE_VALUES else []
        pog_bays = parse_pog_string(pog_raw, logger=None) if pog_raw else []
        all_bays = (lft_bays or []) + (pog_bays or [])
        bay_count = len(all_bays)
        bay_label = f"{bay_count} Bay" if bay_count > 0 else ""

        base_pog_name       = _str(row.get("POG Name", "")) or _make_pog_name(pog_raw, lft_raw)
        base_final_set_name = _str(row.get("Final Set Name", "")) or _make_final_set_name(base_pog_name)
        skus                = stock_sku_seq.get(sid, [])
        sku_set             = tuple(sorted(set(skus)))

        store_info_list.append({
            "store":               sid,
            "bay_label":           bay_label,
            "base_pog_name":       base_pog_name,
            "base_final_set_name": base_final_set_name,
            "skus":                skus,
            "sku_count":           len(skus),
            "sku_set":             sku_set,
            "pog_group_key":       (base_pog_name, sku_set),
            "set_group_key":       (base_final_set_name, sku_set),
            "amt_group_key":       (len(skus), sku_set),
            "full_flow":           ", ".join(skus),
            "preview_flow":        ", ".join(skus[:10]) + ("..." if len(skus) > 10 else ""),
        })

    # 3. Build clusters:
    # A) Set Name clusters: by (base_final_set_name, sku_set)
    set_clusters: Dict[tuple, List[Dict[str, Any]]] = defaultdict(list)
    for info in store_info_list:
        set_clusters[info["set_group_key"]].append(info)

    # B) POG Name clusters: by (base_pog_name, sku_set)
    pog_clusters: Dict[tuple, List[Dict[str, Any]]] = defaultdict(list)
    for info in store_info_list:
        pog_clusters[info["pog_group_key"]].append(info)

    # C) SKU-Count (AMT) clusters: by (sku_count, sku_set)
    amt_clusters: Dict[tuple, List[Dict[str, Any]]] = defaultdict(list)
    for info in store_info_list:
        amt_clusters[info["amt_group_key"]].append(info)

    # 4. Assign versioned POG Names
    pog_to_keys: Dict[str, List[tuple]] = defaultdict(list)
    for k in pog_clusters:
        pog_to_keys[k[0]].append(k)

    versioned_pog_name_map: Dict[tuple, str] = {}
    for base_pog, keys in pog_to_keys.items():
        total_stores = sum(len(pog_clusters[k]) for k in keys)
        if total_stores == 1:
            versioned_pog_name_map[keys[0]] = base_pog
        else:
            keys_sorted = sorted(keys, key=lambda k: -len(pog_clusters[k]))
            for v_num, k in enumerate(keys_sorted, start=1):
                versioned_pog_name_map[k] = f"{base_pog} - V{v_num}"

    # 5. Assign versioned Set Names (Config Group Names)
    set_to_keys: Dict[str, List[tuple]] = defaultdict(list)
    for k in set_clusters:
        set_to_keys[k[0]].append(k)

    versioned_set_name_map: Dict[tuple, str] = {}
    for base_set, keys in set_to_keys.items():
        total_stores = sum(len(set_clusters[k]) for k in keys)
        if total_stores == 1:
            versioned_set_name_map[keys[0]] = base_set
        else:
            keys_sorted = sorted(keys, key=lambda k: -len(set_clusters[k]))
            for v_num, k in enumerate(keys_sorted, start=1):
                versioned_set_name_map[k] = f"{base_set} - V{v_num}"

    # 6. Assign Stock Sku Version & AMT VERSION NAME (Grouped purely by sku_count)
    sku_count_to_keys: Dict[int, List[tuple]] = defaultdict(list)
    for k in amt_clusters:
        sku_count_to_keys[k[0]].append(k)  # k[0] is sku_count

    stock_sku_version_map: Dict[tuple, int] = {}
    amt_version_name_map: Dict[tuple, str] = {}
    for count_val, keys in sku_count_to_keys.items():
        keys_sorted = sorted(keys, key=lambda k: -len(amt_clusters[k]))
        for v_num, k in enumerate(keys_sorted, start=1):
            stock_sku_version_map[k] = v_num
            amt_version_name_map[k] = f"{count_val} SKU Reflow - V{v_num}"

    # 7. Split Set clusters into shared (2+ stores) and unique (1 store) for summary table
    shared_clusters: List[Tuple[tuple, List[Dict[str, Any]]]] = []
    unique_clusters: List[Tuple[tuple, List[Dict[str, Any]]]] = []
    for k, v in set_clusters.items():
        if len(v) >= 2:
            shared_clusters.append((k, v))
        else:
            unique_clusters.append((k, v))

    shared_clusters.sort(key=lambda x: (-len(x[1]), versioned_set_name_map.get(x[0], "")))

    # 8. KPI totals
    total_stores             = len(store_info_list)
    total_unique_configs     = len(set_clusters)
    shared_configs_count     = len(shared_clusters)
    stores_in_shared_count   = sum(len(v) for _, v in shared_clusters)
    stores_with_unique_count = sum(len(v) for _, v in unique_clusters)

    # ── Title (Row 1) ──────────────────────────────────────────────────
    ws["A1"] = "AMT - Identical POG Layout & Stock SKU Flow Analysis"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 24

    # ── Section 1: KPIs (Rows 3–9) ─────────────────────────────────────
    ws["A3"] = "Key Performance Indicators (KPIs)"
    ws["A3"].font = _SECTION_FONT

    kpi_headers = ["Metric", "Value"]
    for c_idx, h in enumerate(kpi_headers, start=1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.font      = _WHITE_FONT_BOLD
        cell.fill      = _NAVY_FILL
        cell.alignment = _ALIGN_LEFT if c_idx == 1 else _ALIGN_RIGHT
        cell.border    = _CELL_BORDER
    ws.row_dimensions[4].height = 20

    kpis = [
        ("Total Stores Analyzed",                    total_stores),
        ("Total Unique Flow Configurations",          total_unique_configs),
        ("Configurations Shared by 2+ Stores",       shared_configs_count),
        ("Stores in Shared Flow Configurations",     stores_in_shared_count),
        ("Stores with Unique Configurations",        stores_with_unique_count),
    ]
    for r_offset, (label, val) in enumerate(kpis, start=5):
        cA = ws.cell(row=r_offset, column=1, value=label)
        cB = ws.cell(row=r_offset, column=2, value=val)
        cA.font = _KPI_LABEL_FONT;  cA.alignment = _ALIGN_LEFT;  cA.border = _CELL_BORDER
        cB.font = _KPI_VAL_FONT;    cB.alignment = _ALIGN_RIGHT; cB.border = _CELL_BORDER
        ws.row_dimensions[r_offset].height = 19

    # ── Section 2: Shared Configurations (Rows 12+) ────────────────────
    ws["A12"] = "Summary of Shared Configurations (Groups with 2+ Stores)"
    ws["A12"].font = _SECTION_FONT

    sum_headers = [
        "Config Group Name",
        "# Stores with Identical Flow",
        "Matching Store Numbers",
        "# Stock SKUs in Flow",
        "Stock SKU Flow Preview (First 10 SKUs)",
    ]
    for c_idx, h in enumerate(sum_headers, start=1):
        cell = ws.cell(row=13, column=c_idx, value=h)
        cell.font      = _WHITE_FONT_BOLD
        cell.fill      = _NAVY_FILL
        cell.alignment = _ALIGN_CENTER if c_idx in (2, 4) else _ALIGN_LEFT
        cell.border    = _CELL_BORDER
    ws.row_dimensions[13].height = 24

    cur_row = 14
    for k, v in shared_clusters:
        cfg_name    = versioned_set_name_map[k]
        store_count = len(v)
        store_nums  = ", ".join(s["store"] for s in v)
        sku_count   = v[0]["sku_count"]
        preview     = v[0]["preview_flow"]

        row_vals = [cfg_name, store_count, store_nums, sku_count, preview]
        for c_idx, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=cur_row, column=c_idx, value=val)
            cell.font      = _DATA_FONT_BOLD if c_idx == 2 else _DATA_FONT
            cell.border    = _CELL_BORDER
            if c_idx == 2:          # "# Stores" highlighted green, centred
                cell.fill      = _LIGHT_GREEN_FILL
                cell.alignment = _ALIGN_CENTER
            elif c_idx == 4:        # "# SKUs" centred
                cell.alignment = _ALIGN_CENTER
            else:
                cell.alignment = _ALIGN_LEFT
        ws.row_dimensions[cur_row].height = 19
        cur_row += 1

    # Spacing before Section 3
    cur_row += 2

    # ── Section 3: Store-by-Store Detail Table ─────────────────────────
    ws.cell(
        row=cur_row, column=1,
        value=f"Store-by-Store Flow Matching Detail (All {total_stores} Stores)",
    ).font = _SECTION_FONT
    cur_row += 1

    detail_headers = [
        "Store",
        "# Stock SKUs",
        "POG Name",
        "# Stores with Identical Flow",
        "Set Name",
        "Stock Sku Version",
        "AMT",
        "AMT VERSION NAME",
        "Bay",
    ]
    hdr_row = cur_row
    for c_idx, h in enumerate(detail_headers, start=1):
        cell           = ws.cell(row=hdr_row, column=c_idx, value=h)
        cell.font      = _WHITE_FONT_BOLD
        cell.fill      = _HD_ORANGE if c_idx <= 5 else _NAVY_FILL
        cell.alignment = _ALIGN_CENTER if c_idx in (1, 2, 4, 6, 9) else _ALIGN_LEFT
        cell.border    = _CELL_BORDER
    ws.row_dimensions[hdr_row].height = 24
    cur_row += 1

    _AMT_GRAY_FILL = PatternFill(fill_type="solid", fgColor="EFEFEF")

    for info in store_info_list:
        set_k = info["set_group_key"]
        pog_k = info["pog_group_key"]
        amt_k = info["amt_group_key"]

        matching_cluster = set_clusters[set_k]
        matching_count   = len(matching_cluster)

        pog_name_val     = versioned_pog_name_map.get(pog_k, info["base_pog_name"])
        set_name_val     = versioned_set_name_map.get(set_k, info["base_final_set_name"])
        sku_version_val  = stock_sku_version_map.get(amt_k, 1)
        amt_val          = f"{info['sku_count']} SKU Reflow"
        amt_ver_name_val = amt_version_name_map.get(amt_k, f"{amt_val} - V{sku_version_val}")

        row_vals = [
            info["store"],
            info["sku_count"],
            pog_name_val,
            matching_count,
            set_name_val,
            sku_version_val,
            amt_val,
            amt_ver_name_val,
            info["bay_label"],
        ]
        for c_idx, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=cur_row, column=c_idx, value=val)
            cell.font      = _DATA_FONT_BOLD if c_idx == 8 else _DATA_FONT
            cell.border    = _CELL_BORDER
            cell.alignment = _ALIGN_CENTER if c_idx in (1, 2, 4, 6, 9) else _ALIGN_LEFT
            if c_idx == 8:
                cell.fill  = _AMT_GRAY_FILL
        ws.row_dimensions[cur_row].height = 19
        cur_row += 1

    # ── Column widths ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # Store
    ws.column_dimensions["B"].width = 16   # # Stock SKUs
    ws.column_dimensions["C"].width = 48   # POG Name
    ws.column_dimensions["D"].width = 26   # # Stores with Identical Flow
    ws.column_dimensions["E"].width = 44   # Set Name
    ws.column_dimensions["F"].width = 18   # Stock Sku Version
    ws.column_dimensions["G"].width = 18   # AMT
    ws.column_dimensions["H"].width = 24   # AMT VERSION NAME
    ws.column_dimensions["I"].width = 12   # Bay



def write_output_bytes(
    input_bytes: bytes,
    planogram_df: pd.DataFrame,
    validation_df: pd.DataFrame,
    wb_data: "WorkbookData | None" = None,
) -> bytes:
    wb = _openpyxl_load(BytesIO(input_bytes))

    # ── 1. Generated Planogram Sheet ───────────────────────────────────────────
    _write_df_to_sheet(wb, OUTPUT_SHEET_PLANOGRAM, planogram_df, fill=_HD_ORANGE)

    # ── 2. Validation Sheet ───────────────────────────────────────────────────
    val_cols = ["Level", "Store", "Bay#", "Message"]
    val_df   = validation_df.reindex(columns=val_cols).fillna("")
    _write_df_to_sheet(wb, OUTPUT_SHEET_VALIDATION, val_df, fill=_ERROR_RED)

    # ── 3. Update 'Store List' Sheet with 'POG Name' (Col B) ───────────────────
    if wb_data is not None:
        store_pog_df = build_store_list_with_pog_name(wb_data)
        _write_df_to_sheet(wb, "Store List", store_pog_df, fill=_NAVY_FILL)

        # ── 4. AMT Executive Dashboard Sheet ──────────────────────────────────
        _write_amt_sheet(wb, store_pog_df, planogram_df, wb_data)

    # ── 5. Reorder Tabs to match expected sequence ─────────────────────────────
    # Expected order: AMT -> Store List -> Stock SKUs and Displays -> Special Order Boards -> Generated Planogram
    desired_tab_order = [
        "AMT",
        "Store List",
        SHEET_STOCK_DISPLAY,
        "Stock SKUs and Displays",
        SHEET_SO,
        "Special Order Boards",
        OUTPUT_SHEET_PLANOGRAM,
        OUTPUT_SHEET_VALIDATION,
    ]
    wb._sheets.sort(
        key=lambda s: desired_tab_order.index(s.title) if s.title in desired_tab_order else 99
    )

    output_bio = BytesIO()
    wb.save(output_bio)
    output_bio.seek(0)
    return output_bio.read()


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 10 — STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<style>
:root {
    --font-primary: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    --color-primary: #8B6F4E;
    --color-primary-hover: #7A6245;
    --color-text: #2B2B2B;
    --color-text-secondary: #212529;
    --color-white: #FFFFFF;
    --color-background: #FAF9F8;
    --color-border-focus: #8B572A;
    --card-bg: #FFFFFF;
    --card-shadow: 0px 1px 5px #AAAAAA;
    --card-radius: 20px;
    --button-radius: 0px;
}

html, body, [class*="css"], [data-testid="stAppViewContainer"], [data-testid="stMain"] {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif !important;
    background-color: #FAF9F8 !important;
    color: #2B2B2B !important;
    font-size: 16px !important;
    line-height: 1.5 !important;
    font-weight: normal !important;
}

[data-testid="stHeader"] {
    background: transparent !important;
}

.main .block-container {
    padding-top: 2rem !important;
    padding-bottom: 2rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 1100px !important;
}

/* ── Typography & Headings ─────────────────────────────────── */
h1, h2, h3, h4, h5, h6 {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif !important;
    color: #2B2B2B !important;
    font-weight: 700 !important;
}

/* ── Hero Banner ───────────────────────────────────────────── */
.hero-banner {
    background: #FFFFFF;
    border: 1px solid #FFFFFF;
    border-radius: 20px;
    box-shadow: 0px 1px 5px #AAAAAA;
    padding: 2.2rem 2.4rem 2rem;
    margin-bottom: 1.8rem;
}
.hero-title {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    font-size: 2.2rem;
    font-weight: 700;
    color: #8B6F4E;
    margin: 0 0 0.4rem 0;
    line-height: 1.2;
}
.hero-sub {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    font-size: 1rem;
    color: #2B2B2B;
    margin: 0;
    line-height: 1.5;
}

/* ── Cards / Panels ────────────────────────────────────────── */
.info-card {
    background: #FFFFFF;
    border: 1px solid #FFFFFF;
    border-radius: 20px;
    box-shadow: 0px 1px 5px #AAAAAA;
    padding: 1.4rem 1.8rem;
    margin-bottom: 1.8rem;
    font-size: 0.95rem;
    color: #2B2B2B;
    line-height: 1.6;
}
.info-card strong {
    color: #8B6F4E;
    font-weight: 600;
}
.info-card code {
    background: #FAF9F8;
    color: #8B572A;
    border: 1px solid #EAE8E4;
    border-radius: 4px;
    padding: 2px 7px;
    font-size: 0.88em;
}

/* ── Stat Cards ────────────────────────────────────────────── */
.stat-card {
    background: #FFFFFF;
    border: 1px solid #FFFFFF;
    border-radius: 20px;
    box-shadow: 0px 1px 5px #AAAAAA;
    padding: 1.4rem 1.6rem;
    text-align: center;
    border-top: 4px solid #8B6F4E;
}
.stat-value {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    font-size: 2.3rem;
    font-weight: 700;
    color: #8B6F4E;
    line-height: 1.1;
}
.stat-label {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    font-size: 0.82rem;
    font-weight: 600;
    color: #212529;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 0.5rem;
}

/* ── Section Headers ───────────────────────────────────────── */
.section-hdr {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
    font-size: 1.05rem;
    font-weight: 700;
    color: #8B6F4E;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    border-bottom: 2px solid #8B6F4E;
    padding-bottom: 0.4rem;
    margin: 1.8rem 0 1.1rem;
}

/* ── File Uploader ─────────────────────────────────────────── */
[data-testid="stFileUploader"] section {
    border: dashed 1px #8B572A !important;
    border-radius: 20px !important;
    background: #FFFFFF !important;
    box-shadow: 0px 1px 5px #AAAAAA !important;
    padding: 1.6rem !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: #8B6F4E !important;
    background: #FCFBF9 !important;
}
[data-testid="stFileUploader"] section button {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif !important;
    background-color: #8B6F4E !important;
    border: 1px solid #8B6F4E !important;
    color: #FFFFFF !important;
    border-radius: 0px !important;
    text-transform: uppercase !important;
    font-size: 90% !important;
    font-weight: 400 !important;
    padding: .375rem .75rem !important;
    height: calc(1.5em + .75em + 2px) !important;
}

/* ── Buttons ───────────────────────────────────────────────── */
.stButton > button,
.stButton > button[kind="primary"],
.stDownloadButton > button {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif !important;
    background-color: #8B6F4E !important;
    border: 1px solid #8B6F4E !important;
    color: #FFFFFF !important;
    border-radius: 0px !important;
    text-transform: uppercase !important;
    font-size: 90% !important;
    font-weight: 400 !important;
    padding: .375rem .75rem !important;
    height: calc(1.5em + .75em + 2px) !important;
    box-shadow: none !important;
    transition: background-color 0.15s ease-in-out, border-color 0.15s ease-in-out !important;
}
.stButton > button:hover,
.stButton > button[kind="primary"]:hover,
.stDownloadButton > button:hover {
    background-color: #7A6245 !important;
    border-color: #7A6245 !important;
    color: #FFFFFF !important;
}

/* ── Expanders ─────────────────────────────────────────────── */
[data-testid="stExpander"] {
    background: #FFFFFF !important;
    border: 1px solid #FFFFFF !important;
    border-radius: 20px !important;
    box-shadow: 0px 1px 5px #AAAAAA !important;
    overflow: hidden !important;
    margin-bottom: 1rem !important;
}
[data-testid="stExpander"] summary {
    font-family: ProximaNova, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif !important;
    color: #2B2B2B !important;
    font-weight: 600 !important;
}

/* ── Dataframes ────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    background: #FFFFFF !important;
    border-radius: 12px !important;
    overflow: hidden !important;
}

/* ── Alerts & Dividers ─────────────────────────────────────── */
hr {
    border-color: #E8E5E0 !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero-banner">
  <div class="hero-title">Home Depot Planogram Generator</div>
  <p class="hero-sub">Upload your Excel workbook to automatically generate a complete bay-level planogram layout for every store.</p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="info-card">
  <strong>Required sheets in workbook:</strong><br>
  &nbsp;&nbsp;• <code>Store List</code> — Store | Current Store POG | Current LFT | Notes<br>
  &nbsp;&nbsp;• <code>Stock SKUs and Displays</code> — Store | Stock SKU | Stock Desc | ... | Facings | Display SKU | Display Desc | ... | Facings | CF<br>
  &nbsp;&nbsp;• <code>Special Order Boards</code> — Store | SO SKU | Description | ... | Facings | CF<br><br>
  <strong>Bay order:</strong> LFT bays are placed <em>first</em>, then POG bays follow.<br>
  <strong>Facings:</strong> Only 1 or 2 are valid — anything else is clamped to 1.<br>
  <strong>Notes rule:</strong> If Notes contains "baja" or "vigo", those SKUs are pushed to the end.
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-hdr">Upload Workbook</div>', unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Choose your Excel workbook (.xlsx)",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

st.markdown("")
run_clicked = st.button("Generate Planogram", type="primary")

if run_clicked:
    if uploaded_file is None:
        st.error("Please upload an Excel workbook first.")
        st.stop()

    file_bytes = uploaded_file.getvalue()
    logger     = PlanogramLogger()

    try:
        with st.spinner("Loading workbook…"):
            wb_data = load_workbook_from_bytes(file_bytes, logger)

        st.markdown("---")
        with st.expander("Column Detection Report — verify before generating", expanded=False):
            st.caption(
                "Shows which actual Excel column header was detected for each logical field. "
                "Facing columns are critical — if they show 'CF', the wrong column was detected."
            )
            col_rows = [
                ("Store List",      "Store ID",            wb_data.cols_sl.get("store", "?")),
                ("Store List",      "Current Store POG",   wb_data.cols_sl.get("pog",   "?")),
                ("Store List",      "Current LFT",         wb_data.cols_sl.get("lft",   "?")),
                ("Store List",      "Notes",               wb_data.cols_sl.get("notes", "?")),
                ("Stock & Display", "Store ID",            wb_data.cols_sd.get("store",      "?")),
                ("Stock & Display", "Stock SKU",           wb_data.cols_sd.get("stock_sku",  "?")),
                ("Stock & Display", "Stock Description",   wb_data.cols_sd.get("stock_desc", "?")),
                ("Stock & Display", "Product Name (exact-match)", wb_data.cols_sd.get("stock_name", "?")),
                ("Stock & Display", "Stock Facing",        wb_data.cols_sd.get("stock_face", "?")),
                ("Stock & Display", "Display SKU",         wb_data.cols_sd.get("disp_sku",   "?")),
                ("Stock & Display", "Display Description", wb_data.cols_sd.get("disp_desc",  "?")),
                ("Stock & Display", "Display Facing",      wb_data.cols_sd.get("disp_face",  "?")),
                ("Stock & Display", "CF",                  wb_data.cols_sd.get("cf",         "?")),
                ("Special Orders",  "Combined Rank",       wb_data.cols_so.get("combined_rank", "?")),
                ("Special Orders",  "Category",            wb_data.cols_so.get("category",      "?")),
                ("Special Orders",  "SKU",                 wb_data.cols_so.get("sku",           "?")),
                ("Special Orders",  "SKU Description",     wb_data.cols_so.get("sku_desc",      "?")),
                ("Special Orders",  "Product Name (exact-match)", wb_data.cols_so.get("name", "?")),
                ("Special Orders",  "OMSID",               wb_data.cols_so.get("omsid",         "?")),
                ("Special Orders",  "CF (Color Flow)",     wb_data.cols_so.get("cf",            "?")),
            ]
            st.dataframe(
                pd.DataFrame(col_rows, columns=["Sheet", "Logical Field", "→ Actual Excel Column Detected"]),
                use_container_width=True,
                hide_index=True,
            )

        with st.spinner("Generating planogram…"):
            planogram_df, validation_df = generate_planogram(wb_data, logger)

        with st.spinner("Writing Excel output…"):
            output_bytes = write_output_bytes(file_bytes, planogram_df, validation_df, wb_data)

    except ValueError as exc:
        st.error(f"Data Error: {exc}")
        st.stop()
    except RuntimeError as exc:
        st.error(f"File Error: {exc}")
        st.stop()
    except Exception as exc:
        st.error(f"Unexpected error: {exc}")
        with st.expander("Technical details"):
            st.code(traceback.format_exc())
        st.stop()

    st.markdown("---")
    st.markdown('<div class="section-hdr">Generation Complete</div>', unsafe_allow_html=True)

    n_stores = planogram_df["Store"].nunique() if not planogram_df.empty else 0
    n_rows   = len(planogram_df)
    n_issues = len(validation_df)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            f'<div class="stat-card"><div class="stat-value">{n_stores:,}</div>'
            f'<div class="stat-label">Stores processed</div></div>',
            unsafe_allow_html=True,
        )
    with c2:
        st.markdown(
            f'<div class="stat-card"><div class="stat-value">{n_rows:,}</div>'
            f'<div class="stat-label">Planogram rows</div></div>',
            unsafe_allow_html=True,
        )
    with c3:
        color = "#8B572A" if n_issues > 0 else "#8B6F4E"
        st.markdown(
            f'<div class="stat-card" style="border-top-color:{color};">'
            f'<div class="stat-value" style="color:{color};">{n_issues:,}</div>'
            f'<div class="stat-label">Validation issues</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("")
    st.download_button(
        label="Download Planogram_Output.xlsx",
        data=output_bytes,
        file_name="Planogram_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    if n_issues > 0:
        st.markdown('<div class="section-hdr">Validation Issues</div>', unsafe_allow_html=True)
        with st.expander(f"Show {n_issues} issue(s)", expanded=True):
            st.dataframe(validation_df, use_container_width=True, hide_index=True)
    else:
        st.success("No validation issues found.")

    if not planogram_df.empty:
        st.markdown('<div class="section-hdr">Planogram Preview (first 200 rows)</div>', unsafe_allow_html=True)
        with st.expander("Show preview", expanded=False):
            st.dataframe(planogram_df.head(200), use_container_width=True, hide_index=True)

st.markdown("---")
st.markdown(
    '<p style="text-align:center;color:#8B6F4E;font-size:0.85rem;font-weight:600;letter-spacing:0.5px;text-transform:uppercase;margin-top:1.5rem;">Home Depot Planogram Generator</p>',
    unsafe_allow_html=True,
)
